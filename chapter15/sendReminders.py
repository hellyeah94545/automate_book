#! /usr/bin/env python3
# sendReminders.py: scripts to send past due email reminders


import sys
import openpyxl
import smtplib


datafile = 'duesRecords.xlsx'
workbook = openpyxl.load_workbook(datafile)
sheet = workbook['Sheet1']

last_column = sheet.max_column
last_row = sheet.max_row
latest_month = sheet.cell(row=1, column=last_column).value

# Check each member's payment status
unpaid_members = {}
for user in range(2, last_row + 1):
    payment = sheet.cell(row=user, column=last_column).value
    if payment != "paid":
        for month in range(2, last_column + 1):
            name = sheet.cell(row=user, column=1).value
            email = sheet.cell(row=user, column=2).value
            unpaid_members[name] = email

# Log in to email account
email_server = "smtp.gmail.com"
sender_email = "hellyeah94545@gmail.com"
sender_password = input("Enter app password: ")
smtpObj = smtplib.SMTP(email_server, 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(sender_email, sender_password)

# Send out reminder emails

for name, email in unpaid_members.items():
    body = "Subject: {0} dues unpaid. \nDear {1}, \n\nRecords show that you have not paid dues for {0}." \
           "Please make this payment as soon as possible. \n\nThank you!".format(latest_month, name)
    send_mail_status = smtpObj.sendmail(sender_email, email, body)
    print(email)
    if send_mail_status != {}:
        print('There was a problem send email to {0}: {1}'.format(email, send_mail_status))
smtpObj.close()

# import textMyself
# textMyself.textmyself("email script run finished")